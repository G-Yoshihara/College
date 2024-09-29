'''Desenvolva um programa que só permita o acesso a pessoas autorizadas (professor, via
autenticação) para posteriormente realizar a média do aluno. Para isto Considere o programa
criado no Fix44.
Incluir uma mensagem na qual deverá aparecer o seu nome, RA e turma antes do resultado final, depois
faça um (print screen) e comentar o(s) resultado(s) do programa após a execução do mesmo.'''


import openpyxl
from openpyxl import Workbook
'''
identificador = "Nome: Guilherme Yoshihara RA: T236JC1"
print(identificador)

# Dicionário com credenciais de acesso
acessopermitido = {"usuario": "GuilhermeY", "senha": "Fix45"}

# Solicitação de nome e senha
acesso = False

while acesso == False:
    usuario = input("Usuário: ")
    senha = input("Senha: ")
    if usuario == acessopermitido["usuario"] and senha == acessopermitido["senha"]:
        print(f"Acesso permitido! Bem-vindo, {usuario}.")
        acesso = True
    else:
        print(f"Acesso negado. Verifique usuário e senha.")
'''

inserir_aluno = True

while inserir_aluno == True:
    inserir_notas = True
    while inserir_notas == True:
    # calcular a média dos alunos
        aluno = input("Digite o nome do aluno: ")
        ra = input("Digite o RA do aluno: ")
        turma = input("Digite a turma do aluno: ")
        np1 = float(input("Digite a primeira nota do aluno: "))
        np2 = float(input("Digite a segunda nota do aluno: "))
        
        mg = ((np1 * 4) + (np2 * 6))/ 10
        
        print(f"Nome: {aluno}, RA: {ra}, Turma: {turma}, Média Final: {mg}")
        
        salvar = str(input("Salvar as informações? (s/n): "))
        if salvar == "s":
            #inserir exportação dos dados para Excel
            workbook = openpyxl.Workbook("Notas_Alunos.xlsx")
            sheet = workbook.active
            data = [
                ["Nome", "RA", "Turma", "Média Final"],
                [aluno, ra, turma, mg]
            ]
            celula1 = sheet.cell(row=1, column=1)
            celula1.value = "Nome"
            celula2 = sheet.cell(row=1, column=2)
            celula2.value = "RA"
            celula3 = sheet.cell(row=1, column=3)
            celula3.value = "Turma"
            celula4 = sheet.cell(row=1, column=4)
            celula4.value = "Média Final"

            continuar = str(input("Deseja inserir novos alunos? (s/n): "))
            if continuar == "n":
                inserir_aluno = False
                inserir_notas = False
            else:
                print("Insira os dados do novo aluno:")
        else:
            print("Revise as informações")

workbook.save("Notas_Alunos.xlsx")
print("As informações foram salvas no arquivo Excel. Até a próxima.")